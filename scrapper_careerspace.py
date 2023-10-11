import json
import random
import re
import requests
import shutil
import time

from bs4 import BeautifulSoup

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from pathlib import Path

from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions

from typing import Pattern, Union

from core_utils.config_dto import ConfigDTO
from core_utils.constants import (ASSETS_PATH, CRAWLER_CONFIG_PATH,
                                  NUM_ARTICLES_UPPER_LIMIT,
                                  TIMEOUT_LOWER_LIMIT, TIMEOUT_UPPER_LIMIT)
from core_utils.vacancy import Vacancy
from core_utils.visualizer import visualize


class IncorrectSeedURLError(Exception):
    """"
    Seed URLs value is not of the list type or
    its elements either do not match the standard URL pattern
    or not of the string type
    """


class NumberOfArticlesOutOfRangeError(Exception):
    """
    The number of articles that must be parsed is out of given range
    """


class IncorrectNumberOfArticlesError(Exception):
    """
    The number of articles that must be parsed is not integer
    """


class IncorrectHeadersError(Exception):
    """
    Headers value is not of the dictionary type
    """


class IncorrectEncodingError(Exception):
    """
    Encoding value is not of a string type
    """


class IncorrectTimeoutError(Exception):
    """
    Timeout value is out of range from 1 to the given value
    """


class IncorrectVerifyError(Exception):
    """
    Should_verify_certificate value is not of a boolean type
    """


class Config:
    """
    Unpacks and validates configurations
    """

    seed_urls: list[str]
    num_articles: int
    headers: dict[str, str]
    encoding: str
    timeout: int
    verify_certificate: bool
    headless_mode: bool

    def __init__(self, path_to_config: Path) -> None:
        """
        Initializes an instance of the Config class
        """
        self.path_to_config = path_to_config
        self._validate_config_content()

        config_dto = self._extract_config_content()
        self._seed_urls = config_dto.seed_urls
        self._num_articles = config_dto.total_articles
        self._headers = config_dto.headers
        self._encoding = config_dto.encoding
        self._timeout = config_dto.timeout
        self._should_verify_certificate = config_dto.should_verify_certificate
        self._headless_mode = config_dto.headless_mode

    def _extract_config_content(self) -> ConfigDTO:
        """
        Returns config values
        """
        with open(self.path_to_config, 'r', encoding='utf-8') as f:
            config_params = json.load(f)
        return ConfigDTO(**config_params)

    def _validate_config_content(self) -> None:
        """
        Ensure configuration parameters
        are not corrupt
        """
        config_dto = self._extract_config_content()

        if not isinstance(config_dto.seed_urls, list):
            raise IncorrectSeedURLError('seed URLs must be of passed as a list')

        if not all(re.match('^https?://[w{3}]?.*/', url) and
                   isinstance(url, str) for url in config_dto.seed_urls):
            raise IncorrectSeedURLError('seed URLs either do not match the standard pattern '
                                        'or not strings')

        if not isinstance(config_dto.total_articles, int) or \
                isinstance(config_dto.total_articles, bool) \
                or config_dto.total_articles < 1:
            raise IncorrectNumberOfArticlesError('total number of articles to parse is not integer')

        if NUM_ARTICLES_UPPER_LIMIT and config_dto.total_articles > NUM_ARTICLES_UPPER_LIMIT:
            raise NumberOfArticlesOutOfRangeError('total number of articles is '
                                                  'out of range from 1 to the given value')

        if not isinstance(config_dto.headers, dict):
            raise IncorrectHeadersError('headers are not in a form of a dictionary')

        if not isinstance(config_dto.encoding, str):
            raise IncorrectEncodingError('encoding must be specified as a string')

        if config_dto.timeout not in range(TIMEOUT_LOWER_LIMIT, TIMEOUT_UPPER_LIMIT + 1):
            raise IncorrectTimeoutError('timeout value must be a positive integer '
                                        'less than the given value')

        if not isinstance(config_dto.should_verify_certificate, bool) or \
                not isinstance(config_dto.headless_mode, bool):
            raise IncorrectVerifyError('verify certificate value must either be True or False')

    def get_seed_urls(self) -> list[str]:
        """
        Retrieve seed urls
        """
        return self._seed_urls

    def get_num_articles(self) -> int:
        """
        Retrieve total number of articles to scrape
        """
        return self._num_articles

    def get_headers(self) -> dict[str, str]:
        """
        Retrieve headers to use during requesting
        """
        return self._headers

    def get_encoding(self) -> str:
        """
        Retrieve encoding to use during parsing
        """
        return self._encoding

    def get_timeout(self) -> int:
        """
        Retrieve number of seconds to wait for response
        """
        return self._timeout

    def get_verify_certificate(self) -> bool:
        """
        Retrieve whether to verify certificate
        """
        return self._should_verify_certificate

    def get_headless_mode(self) -> bool:
        """
        Retrieve whether to use headless mode
        """
        return self._headless_mode


def make_request(url: str, config: Config) -> requests.models.Response:
    """
    Delivers a response from a request
    with given configuration
    """
    time.sleep(random.randint(2, 5))
    response = requests.get(url,
                            headers=config.get_headers(),
                            timeout=config.get_timeout(),
                            verify=config.get_verify_certificate())
    response.raise_for_status()
    response.encoding = config.get_encoding()
    return response


class Crawler:
    """
    Crawler implementation
    """

    url_pattern: Union[Pattern, str]

    def __init__(self, config: Config) -> None:
        """
        Initializes an instance of the Crawler class
        """
        self.config = config
        self.urls = []

    def _extract_url(self, article_bs: BeautifulSoup) -> str:
        """
        Finds and retrieves URL from HTML
        """
        url = article_bs.get('href')
        if url and url.count('/') == 2 and url[:5] == '/job/':
            return 'https://careerspace.app' + str(url)
        return ''

    def find_articles(self, response) -> None:
        """
        Finds articles on a current page level
        """
        article_bs = BeautifulSoup(response, 'lxml')
        links = article_bs.find_all('a', {'class': 'job-card__i'})
        for link in links:
            if len(self.urls) < self.config.get_num_articles():
                url = self._extract_url(link)
                if url and url not in self.urls:
                    self.urls.append(url)

    def scroll_site(self) -> None:
        chrome_options = ChromeOptions()
        chrome_options.add_argument('--start-maximized')
        if self.config.get_headless_mode():
            chrome_options.add_argument('--headless=new')

        driver = webdriver.Chrome(options=chrome_options)
        driver.get(self.config.get_seed_urls()[0])

        scroll_pause_time = 2
        last_height = driver.execute_script("return document.body.scrollHeight")
        while len(self.urls) < self.config.get_num_articles():
            driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
            current_html = driver.page_source
            self.find_articles(current_html)

            time.sleep(scroll_pause_time)
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height
        driver.quit()

    def get_search_urls(self) -> list:
        """
        Returns seed_urls param
        """
        return self.config.get_seed_urls()


class HTMLParser:
    """
    VacancyParser implementation
    """
    def __init__(self, full_url: str, vacancy_id: int, config: Config) -> None:
        """
        Initializes an instance of the HTMLParser class
        """
        self.full_url = full_url
        self.vacancy_id = vacancy_id
        self.config = config
        self.vacancy = Vacancy(url=self.full_url, vacancy_id=self.vacancy_id)

    def _fill_vacancy_info(self, vacancy_soup: BeautifulSoup) -> None:
        """
        Extracts vacancy information from its page
        """
        upper = vacancy_soup.find('div', {'class': 'j-d-h__inner'})

        self.vacancy.name = upper.find('h3').text

        self.vacancy.level = self.identify_level(self.vacancy.name)

        location = self.extract_location(upper.find_all('span', {'class': 'job-lb__tx'}))
        self.vacancy.location = location[0]
        self.vacancy.remote = location[1]
        self.vacancy.hybrid = location[2]

        employer = vacancy_soup.find('div', {'class': 'j-d-h__company cs-df-alc'})
        if employer:
            self.vacancy.employer = employer.text.strip()
        else:
            employer = vacancy_soup.find('div', {'class': 'j-d-cm__name'})
            if employer:
                self.vacancy.employer = employer.text.strip()
            else:
                self.vacancy.employer = ''

        lower = vacancy_soup.find('div', {'class': 'j-d__content'})

        salary = self.extract_salary(lower.find('span', {'class': 'price'}).text)
        self.vacancy.salary_bottom_line = salary[0]
        self.vacancy.salary_top_line = salary[1]

    def extract_location(self, location) -> tuple:
        city = None
        remote = False
        hybrid = False
        location = [l.text.strip() for l in location]
        if 'Удаленно' in location:
            remote = True
            location.remove('Удаленно')
        if 'Гибрид' in location:
            hybrid = True
            location.remove('Гибрид')
        if location:
            city = location[0]

        return city, remote, hybrid

    def extract_salary(self, salary) -> tuple:
        if 'от' in salary:
            bottom = int(re.sub(r'\D', '', salary))
            top = None

        elif 'до' in salary:
            bottom = None
            top = int(re.sub(r'\D', '', salary))

        else:
            salary_info = salary.split(' - ')
            if len(salary_info) == 1:
                bottom = top = int(re.sub(r'\D', '', salary_info[0]))
            else:
                bottom, top = int(re.sub(r'\D', '', salary_info[0])), int(re.sub(r'\D', '', salary_info[1]))

        return bottom, top

    def identify_level(self, name) -> list:
        """
        Extracts qualification level from the vacancy name
        """
        levels = []
        clean_name = name.strip().lower()
        level_names = {
            'Junior': ['junior', 'младший', 'стажер', 'стажёр'],
            'Middle': ['middle', 'средний'],
            'Senior': ['senior', 'старший', 'lead', 'ведущий']
                       }

        for level, possible_names in level_names.items():
            if any(n in clean_name for n in possible_names):
                levels.append(level)

        return ', '.join(levels) if levels else ''

    def parse(self) -> Union[Vacancy, bool, list]:
        """
        Parses each vacancy
        """
        response = make_request(self.full_url, self.config)
        vacancy_soup = BeautifulSoup(response.text, 'lxml')
        self._fill_vacancy_info(vacancy_soup)
        return self.vacancy


def prepare_environment(base_path: Union[Path, str]) -> None:
    """
    Creates ASSETS_PATH folder if no created and removes existing folder
    """
    if base_path.exists():
        shutil.rmtree(base_path)
    base_path.mkdir(parents=True)


def style_excel(ws: Worksheet) -> None:
    """
    Styles excel worksheet
    """
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        adjusted_width = (max_length + 1) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    for cell in ws['1:1']:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='BCB7B6', fill_type="solid")


def main() -> None:
    """
    Entrypoint for scrapper module
    """
    configuration = Config(path_to_config=CRAWLER_CONFIG_PATH)
    prepare_environment(ASSETS_PATH)
    crawler = Crawler(config=configuration)
    crawler.scroll_site()

    print('urls have been collected')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Careerspace'
    excel_path = ASSETS_PATH / 'job_offers.xlsx'
    wb.save(excel_path)
    headings = list(Vacancy('', 0).get_info().keys())[1:]
    ws.append(headings)

    for idx, url in enumerate(crawler.urls, start=1):
        parser = HTMLParser(full_url=url, vacancy_id=idx, config=configuration)
        try:
            parsed_vacancy = parser.parse()
        except (requests.exceptions.ConnectTimeout,
                requests.exceptions.ConnectionError,
                requests.exceptions.ReadTimeout,
                requests.exceptions.HTTPError):
            continue
        else:
            if isinstance(parsed_vacancy, Vacancy):
                row = list(parsed_vacancy.get_info().values())[1:]
                ws.append(row)
                wb.save(excel_path)
    print('urls have been parsed')
    style_excel(ws)
    wb.save(excel_path)
    wb = load_workbook(excel_path)
    ws = wb.worksheets[0]
    statistics_path = ASSETS_PATH / 'statistics_careerspace.html'
    visualize(ws, statistics_path)


if __name__ == "__main__":
    main()
